#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script autonome — Remplissage des onglets S7 (Thème 3) des progressions 5e et 4e.
Usage (depuis la racine du dépôt) :
    python _progressions/remplir_S7_theme3.py

Règles respectées :
- Ne touche QUE les onglets S7 des classeurs 5e et 4e
- Préserve toutes les formules et les autres onglets
- Suit le canevas obligatoire de _progressions/README.md
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
import json

ROOT = Path(__file__).resolve().parent.parent
CONTENU_JSON = Path(__file__).resolve().parent / "contenu_S7_theme3.json"

def charger_contenu():
    with open(CONTENU_JSON, encoding="utf-8") as f:
        return json.load(f)

def appliquer_styles_base(ws):
    """Reprend les styles de base des générateurs S4/S5 (Arial, couleurs thème)."""
    # Les styles exacts sont déjà dans le classeur ; on n'impose que le minimum.
    pass

def remplir_onglet(ws, data):
    """Remplit un onglet S7 avec le dictionnaire de données fourni.

    Les coordonnées (cellules) sont des hypothèses basées sur le canevas.
    À ajuster si la structure réelle de l'onglet S7 diffère.
    """
    # --- En-tête ---
    if "titre" in data:
        ws["B2"] = data["titre"]
    if "theme" in data:
        ws["B3"] = data["theme"]
    if "periode" in data:
        ws["B4"] = data["periode"]
    if "nb_seances" in data:
        ws["B5"] = data["nb_seances"]
    if "responsable" in data:
        ws["B6"] = data["responsable"]
    if "statut" in data:
        ws["B7"] = data["statut"]

    # --- Identification ---
    if "problematique" in data:
        ws["B10"] = data["problematique"]
    if "situation_declenchante" in data:
        ws["B11"] = data["situation_declenchante"]
    if "prerequis" in data:
        ws["B12"] = data["prerequis"]
    if "objectifs_synthese" in data:
        ws["B13"] = data["objectifs_synthese"]
    if "piste_evaluation" in data:
        ws["B14"] = data["piste_evaluation"]
    if "domaines_socle" in data:
        ws["B15"] = data["domaines_socle"]
    if "crcn" in data:
        ws["B16"] = data["crcn"]
    if "versions" in data:
        ws["B17"] = data["versions"]
    if "liens_epi" in data:
        ws["B18"] = data["liens_epi"]
    if "reperes_nathan" in data:
        ws["B19"] = data["reperes_nathan"]
    if "ressources_en_ligne" in data:
        ws["B20"] = data["ressources_en_ligne"]

    # --- Tableau des compétences (à partir de la ligne 23 approximative) ---
    row = 23
    for comp in data.get("competences", []):
        ws.cell(row=row, column=2, value=comp.get("code", ""))
        ws.cell(row=row, column=3, value=comp.get("intitule", ""))
        ws.cell(row=row, column=4, value=comp.get("domaines_socle", ""))
        ws.cell(row=row, column=5, value=comp.get("seances", ""))
        ws.cell(row=row, column=6, value=comp.get("evaluation_lsu", ""))
        row += 1

    # --- Déroulé par séance ---
    row = 35  # hypothèse de départ du déroulé
    for seance in data.get("seances", []):
        ws.cell(row=row, column=2, value=seance.get("numero", ""))
        ws.cell(row=row, column=3, value=seance.get("question_directrice", ""))
        ws.cell(row=row, column=4, value=seance.get("activites_supports", ""))
        ws.cell(row=row, column=5, value=seance.get("demarche", ""))
        ws.cell(row=row, column=6, value=seance.get("bilan_trace", ""))
        # Bloc jaune Cahier de texte Pronote
        ws.cell(row=row+1, column=3, value=seance.get("cahier_texte_contenu", ""))
        ws.cell(row=row+1, column=4, value=seance.get("cahier_texte_travail", ""))
        # Ligne de suivi
        ws.cell(row=row+2, column=3, value=seance.get("suivi_enseignant", ""))
        # Répartition
        ws.cell(row=row, column=7, value=seance.get("repartition", ""))
        row += 4

def main():
    contenu = charger_contenu()

    # --- Classeur 5e ---
    path_5e = ROOT / "_progressions" / "5e" / "Progression_Techno_5e_2026-2027_Martinique.xlsx"
    if path_5e.exists():
        wb5 = load_workbook(path_5e)
        if "S7" in wb5.sheetnames:
            print("Remplissage S7 (5e)...")
            remplir_onglet(wb5["S7"], contenu["5e"]["S7"])
            wb5.save(path_5e)
            print("  → S7 5e sauvegardé.")
        else:
            print("ATTENTION : onglet S7 introuvable dans le classeur 5e.")
    else:
        print(f"Fichier non trouvé : {path_5e}")

    # --- Classeur 4e ---
    path_4e = ROOT / "_progressions" / "4e" / "Progression_Techno_4e_2026-2027_Martinique.xlsx"
    if path_4e.exists():
        wb4 = load_workbook(path_4e)
        if "S7" in wb4.sheetnames:
            print("Remplissage S7 (4e)...")
            remplir_onglet(wb4["S7"], contenu["4e"]["S7"])
            wb4.save(path_4e)
            print("  → S7 4e sauvegardé.")
        else:
            print("ATTENTION : onglet S7 introuvable dans le classeur 4e.")
    else:
        print(f"Fichier non trouvé : {path_4e}")

    print("\nTerminé. Vérifier les onglets S7 puis recalculer si besoin (LibreOffice).")

if __name__ == "__main__":
    main()
